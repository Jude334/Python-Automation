if __name__ == '__main__':
    
    import multiprocessing
    import cpuinfo
    import os
    from psutil import virtual_memory
    import psutil
    import math
    import openpyxl
    
    #Call freeze_support or Pyinstaller will recursively fork infinite processes forever
    multiprocessing.freeze_support()
    computername = os.environ['COMPUTERNAME']
    user = os.getlogin()
    ram = virtual_memory()
    processor = cpuinfo.get_cpu_info()['brand']
    cores = psutil.cpu_count(logical=False)
    disk = psutil.disk_usage('/')

    #prints out formatted info
    print('\nComputer name: ' + computername)
    print('\nUser: ' + user)
    print('\nTotal RAM:')
    print(ram.total / (1024.0 ** 3),'GB')
    print('\nProcessor: ' + processor)
    print('\nCores: ' + str(cores))
    print('\nFree space:')
    print(disk.free / (1024.0 ** 3),'GB')
    print('\nTotal space:')
    print(disk.total / (1024.0 ** 3),'GB')

    #Make Excel Workbook object
    file = '//path/to/excelsheet'
    wb = openpyxl.load_workbook(file)
    sheet = wb['Sheet1']
    #Get number of rows
    rows = sheet.max_row
    #Add each item to array
    arr = []
    for i in range(2, rows+1):
        arr.append(sheet.cell(row=i, column=1).value)

    if computername in arr:
        print('\nThis computer is already logged in Excel Sheet.')
    else:
        #Add new info to new row
        rows += 1
        sheet['A' + str(rows)] = computername
        sheet['B' + str(rows)] = user
        sheet['C' + str(rows)] = str(math.floor(float(ram.total / (1024.0 ** 3)))) + 'GB'
        sheet['D' + str(rows)] = processor + ' | ' + str(cores) + ' cores'
        sheet['E' + str(rows)] = str(math.floor(float(disk.free / (1024.0 ** 3)))) + 'GB/' + str(math.floor(float(disk.total / (1024.0 ** 3)))) + 'GB'

        wb.save(file)

        print('\nComputer info logged in Excel Sheet.')

    input('\nPress any key to exit')
